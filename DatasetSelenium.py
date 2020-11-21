## VERİ ÇEKME KISMI 
# Veri çekme için gerekli kütüphanelerin import edilmesi
from selenium import webdriver
from  openpyxl import *

 # çekilen verinin yazılacağı excell dosyasının oluşturulması
sahibinden = Workbook()

islenmemis = sahibinden.active
islenmemis.title="Önişleme öncesi"
islenmemis.append(["Alan","Oda Sayısı","Fiyat"])

# Chrome web driverinın çalıştırılması

browser = webdriver.Chrome("C://Users//Burak//Desktop//chromedriver")


# web sitesinden veri çekecek döngü
for k in range(0,601,50):
    browser.get(f"https://www.sahibinden.com/satilik-daire/istanbul-basaksehir-basaksehir-basaksehir-mh.?pagingOffset={k}&pagingSize=50&address_region=1&sorting=date_desc&a20=38470&a20=38474&a20=38471")

    for i in range(1,52,1):
        if i == 5 or i == 4:
            i +=1
            continue
        
        # siteden kategorik verilerin çekilmesi
      
        metrekare=browser.find_element_by_xpath(f"//*[@id='searchResultsTable']/tbody/tr[{i}]/td[3]").text
        odasayisi=browser.find_element_by_xpath(f"//*[@id='searchResultsTable']/tbody/tr[{i}]/td[4]").text
        fiyat=browser.find_element_by_xpath(f"//*[@id='searchResultsTable']/tbody/tr[{i}]/td[5]").text
        fiyat=fiyat.replace(" TL","")
        fiyat=fiyat.replace(".","")
        
          # çekilen işlenmemiş verilerin excell sayfasına kaydedilmesi
        islenmemis.append([metrekare,odasayisi,fiyat])
        
        
#kaydedilen dosyanın lokalde oluştrulup kapatılması
sahibinden.save("evlerim.xlsx")
sahibinden.close()
browser.quit()


## VERİ ÖN İŞLEME KISMI
import pandas as pd
from sklearn.preprocessing import LabelEncoder,OneHotEncoder

#Verisetinin okunması.
dataset = pd.read_excel("evlerim.xlsx")

# Fiyat stringlerini int veriye çeviriyoruz
fiyat=pd.DataFrame(data=dataset["Fiyat"])
fiyat = fiyat.astype(int)
dataset["Fiyat"]=fiyat

# Alan stringlerini int veriye çeviriyoruz
alan=pd.DataFrame(data=dataset["Alan"])
alan = alan.astype(int)
dataset["Alan"]=alan


#oda sütunundaki verilerin seçilmesi
oda = dataset.iloc[:,1:2].values


#oda sayısı stringlerinin sayısal verilere dönüştürülmesi.
lblencoder= LabelEncoder()
oda[:,0]=lblencoder.fit_transform(oda[:,0])


#sayısala çevirdiğimiz veriye onehotencoder ile kategorik verilere dönüştürülmesi
onehotencoder = OneHotEncoder()
odaonehot = onehotencoder.fit_transform(oda).toarray()

# Önişleme yaptığımız verileri yazmak için yeni bir excell sayfası oluşturulması
dosya=load_workbook("evlerim.xlsx")
sayfa2=dosya.create_sheet("Önişleme Sonrası")
sayfa2.append(['Oda1','Oda2','Oda3','Alan','Fiyat'])

# Önişlenmiş verilerimizi açtığımız excell sayfasına yazılması
for k in range(0,len(odaonehot)):
    sayfa2.append([odaonehot[k,0],odaonehot[k,1],odaonehot[k,2],dataset["Alan"][k],dataset["Fiyat"][k]]) 

# Dosyanın kaydedilip kapatılması
dosya.save("evlerim.xlsx")
dosya.close()







